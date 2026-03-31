import os
import tempfile
import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Pipeline, SalesLead, Task, User, disable_metrics_events, pipeline_support


class FilterAndExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        db_path = os.path.join(cls.temp_dir.name, 'test_bitcrm.db')

        class TestConfig:
            TESTING = True
            SECRET_KEY = 'test-secret'
            SQLALCHEMY_DATABASE_URI = f'sqlite:///{db_path}'
            SQLALCHEMY_TRACK_MODIFICATIONS = False
            SQLALCHEMY_ENGINE_OPTIONS = {}
            CACHE_TYPE = 'SimpleCache'
            UPLOAD_FOLDER = os.path.join(cls.temp_dir.name, 'uploads')
            EXCEL_TEMPLATES_FOLDER = os.path.join(cls.temp_dir.name, 'templates')

        cls.app = create_app(TestConfig)
        cls.ctx = cls.app.app_context()
        cls.ctx.push()

        db.session.execute(pipeline_support.delete())
        Task.query.delete()
        Pipeline.query.delete()
        SalesLead.query.delete()
        User.query.delete()
        db.session.commit()

        cls.bruce = User(username='Bruce', email='bruce@test.local', role='admin')
        cls.bruce.set_password('bitcrm')
        cls.bruce.set_column_preferences('leads', ['company', 'owner'])
        cls.bruce.set_column_preferences('pipeline', ['company', 'stage', 'tcv_usd'])

        cls.anthony = User(username='Anthony', email='anthony@test.local', role='sales')
        cls.anthony.set_password('bitcrm')

        db.session.add_all([cls.bruce, cls.anthony])
        db.session.commit()

        cls.bruce_id = cls.bruce.id
        cls.anthony_id = cls.anthony.id

        with disable_metrics_events():
            db.session.add_all([
                SalesLead(
                    name='Lead Bruce',
                    company='Acme China',
                    leads_status='Qualified',
                    source='Website',
                    owner_id=cls.bruce_id,
                    date_added=date(2026, 3, 1),
                ),
                SalesLead(
                    name='Lead Anthony',
                    company='Acme Korea',
                    leads_status='Waiting to be Contacted',
                    source='Referral',
                    owner_id=cls.anthony_id,
                    date_added=date(2026, 3, 2),
                ),
                SalesLead(
                    name='Lead Other',
                    company='Globex',
                    leads_status='Qualified',
                    source='Website',
                    owner_id=cls.anthony_id,
                    date_added=date(2026, 3, 3),
                ),
                Pipeline(
                    name='Contact Bruce',
                    company='Acme Telecom',
                    owner_id=cls.bruce_id,
                    product='Cloud Voice',
                    mrc_usd=100,
                    otc_usd=0,
                    contract_term_yrs=1,
                    gp_margin=0.5,
                    win_rate=0.5,
                    stage='1) Prospecting',
                    level='Committed',
                    date_added=date(2026, 3, 1),
                ),
                Pipeline(
                    name='Contact Anthony',
                    company='Acme Networks',
                    owner_id=cls.anthony_id,
                    product='Security',
                    mrc_usd=100,
                    otc_usd=0,
                    contract_term_yrs=2,
                    gp_margin=0.6,
                    win_rate=0.7,
                    stage='2) Lead Qualified',
                    level='Stretch',
                    date_added=date(2026, 3, 2),
                ),
                Pipeline(
                    name='Contact Other',
                    company='Globex Systems',
                    owner_id=cls.anthony_id,
                    product='Connectivity',
                    mrc_usd=100,
                    otc_usd=0,
                    contract_term_yrs=1,
                    gp_margin=0.4,
                    win_rate=0.3,
                    stage='6b) Deal Lost',
                    level='Stretch',
                    date_added=date(2026, 3, 3),
                ),
            ])
            db.session.commit()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        cls.ctx.pop()
        cls.temp_dir.cleanup()

    def setUp(self):
        self.client = self.app.test_client()
        response = self.client.post(
            '/login',
            data={'username': 'Bruce', 'password': 'bitcrm'},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)

    def _read_workbook(self, response):
        workbook = load_workbook(BytesIO(response.data))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        return headers, rows

    def test_leads_company_filter_and_owner_can_reset_to_all(self):
        filtered = self.client.get(
            f'/leads/?company=Acme&owner={self.anthony_id}&sort=date_added&order=desc'
        )
        self.assertEqual(filtered.status_code, 200)
        filtered_html = filtered.get_data(as_text=True)
        self.assertIn('Acme Korea', filtered_html)
        self.assertNotIn('Acme China', filtered_html)

        reset = self.client.get('/leads/?company=Acme&sort=date_added&order=desc')
        self.assertEqual(reset.status_code, 200)
        reset_html = reset.get_data(as_text=True)
        self.assertIn('Acme Korea', reset_html)
        self.assertIn('Acme China', reset_html)
        self.assertLess(reset_html.index('name="company"'), reset_html.index('All Statuses'))

    def test_leads_export_matches_visible_columns_and_filters(self):
        response = self.client.get(
            f'/leads/export?company=Acme&owner={self.bruce_id}&sort=date_added&order=desc'
        )
        self.assertEqual(response.status_code, 200)

        headers, rows = self._read_workbook(response)
        self.assertEqual(headers, ['Company', 'Owner'])
        self.assertEqual(rows, [('Acme China', 'Bruce')])

    def test_pipeline_company_filter_and_export_match_page_state(self):
        page = self.client.get('/pipeline/?company=Acme&sort=date_added&order=desc')
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('Acme Telecom', html)
        self.assertIn('Acme Networks', html)
        self.assertNotIn('Globex Systems', html)
        self.assertLess(html.index('name="company"'), html.index('Sign Date'))

        export = self.client.get('/pipeline/export?company=Acme&sort=date_added&order=desc')
        self.assertEqual(export.status_code, 200)

        headers, rows = self._read_workbook(export)
        self.assertEqual(headers, ['Company', 'Stage', 'TCV'])
        self.assertEqual(
            rows,
            [
                ('Acme Networks', '2) Lead Qualified', 2400),
                ('Acme Telecom', '1) Prospecting', 1200),
            ],
        )


if __name__ == '__main__':
    unittest.main()
